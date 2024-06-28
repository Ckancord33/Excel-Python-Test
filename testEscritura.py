#Importar los diferentes módulos necesarios
from openpyxl import Workbook
from openpyxl.styles import Font
import time

#Crear un libro de trabajo donde se manejarán las hojas de cálculo
book = Workbook()
sheet = book.active

#Escribir en las celdas especificas de la hoja de cálculo
sheet['A1'] = "Hello"
sheet['A2'] = "World"

#Como diseñar la celda
sheet["B1"] = "Rango"
sheet["B1"].font = Font(color="FF0000", bold=True, italic=True)

#Como modificar varias celdas con un for
for i in range(2, 14):
    sheet[f"B{i}"] = i**2

#Crear una nueva hoja de cálculo
sheet2 = book.create_sheet("hoja_2")
sheet2["A1"] = "Hola hoja 2"
sheet2["A2"] = time.strftime("%x")

#Uniones de celdas
sheet3 = book.create_sheet("hoja_3")
sheet3.merge_cells("A1:D1")
sheet3["A1"] = "Unión de celdas"




book.save("ExcelTest.xlsx")